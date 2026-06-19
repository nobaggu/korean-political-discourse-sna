"""
Step 2 (BigKinds 버전): BigKinds 엑셀 → 본문 크롤링 → articles_with_text_all.json

흐름:
  1. BigKinds 엑셀 읽기 (559건, 2국가론)
  2. 2024~2025 기사만 필터 (2026 제외)
  3. BigKinds 인물 필드로 우리 정치인 사전 필터
  4. URL로 본문 크롤링
  5. articles_with_text_all.json 저장 (03b와 동일 포맷)
"""

import json
import os
import re
import time
import requests
import openpyxl
from bs4 import BeautifulSoup
from tqdm import tqdm

os.makedirs("data/raw", exist_ok=True)

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    )
}

# 수집 기간
DATE_FROM = "20240101"
DATE_TO   = "20260618"


def load_assembly():
    with open("data/assembly_members.json", "r", encoding="utf-8") as f:
        return json.load(f)


def crawl_article(url: str) -> str | None:
    try:
        resp = requests.get(url, headers=HEADERS, timeout=10)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")

        # 네이버 뉴스
        if "news.naver.com" in url:
            for sel in ["#dic_area", "#articleBodyContents", ".go_trans._article_content", "#newsct_article"]:
                tag = soup.select_one(sel)
                if tag:
                    return tag.get_text(separator="\n", strip=True)

        # 일반 언론사
        for sel in ["article", "main", ".article-body", ".article_body", "#article-view-content-div",
                    ".news_body", ".article-text", "#articleBody"]:
            tag = soup.select_one(sel)
            if tag and len(tag.get_text(strip=True)) > 200:
                return tag.get_text(separator="\n", strip=True)

        # p 태그 합치기 (fallback)
        paragraphs = soup.find_all("p")
        text = "\n".join(p.get_text(strip=True) for p in paragraphs if len(p.get_text(strip=True)) > 30)
        return text if len(text) > 200 else None

    except Exception:
        return None


def read_bigkinds_excel(path: str) -> list:
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    rows = []
    for row in range(2, ws.max_row + 1):
        rows.append({
            "id":        ws.cell(row, 1).value,
            "date":      str(ws.cell(row, 2).value or ""),
            "press":     ws.cell(row, 3).value,
            "title":     ws.cell(row, 5).value,
            "persons":   str(ws.cell(row, 12).value or ""),   # BigKinds 인물 추출
            "keywords":  str(ws.cell(row, 15).value or ""),
            "body":      str(ws.cell(row, 17).value or ""),   # 200자 요약 (참고용)
            "url":       str(ws.cell(row, 18).value or ""),
        })
    return rows


def crawl_from_bigkinds(excel_path: str):
    assembly = load_assembly()
    assembly_names = list(assembly.keys())

    print(f"BigKinds 엑셀 읽는 중...")
    articles = read_bigkinds_excel(excel_path)
    print(f"  총 {len(articles)}건 로드")

    # 1. 날짜 필터 (2024~2025)
    date_filtered = [
        a for a in articles
        if DATE_FROM <= a["date"] <= DATE_TO
    ]
    print(f"  날짜 필터 후: {len(date_filtered)}건 (2024.01~2025.03)")

    # 2. BigKinds 인물 필드로 사전 필터
    #    인물 필드에 우리 313명 중 1명 이상 언급된 기사만
    pol_filtered = []
    for a in date_filtered:
        persons_in_article = [p for p in assembly_names if p in a["persons"] or p in (a["title"] or "")]
        if persons_in_article:
            a["politicians_pre"] = persons_in_article
            pol_filtered.append(a)

    print(f"  정치인 등장 기사: {len(pol_filtered)}건")

    # 3. URL 크롤링
    results = []
    failed  = 0

    for article in tqdm(pol_filtered, desc="본문 크롤링 중"):
        url = article["url"].strip()
        if not url or url == "None":
            failed += 1
            continue

        full_text = crawl_article(url)
        if not full_text:
            failed += 1
            time.sleep(0.3)
            continue

        # 본문에서 정치인 재확인
        politicians_found = [p for p in assembly_names if p in full_text]
        if not politicians_found:
            time.sleep(0.3)
            continue

        results.append({
            "keyword":          "2국가론",
            "title":            article["title"],
            "description":      article["body"],
            "link":             url,
            "originallink":     url,
            "pub_date":         f"{article['date'][:4]}-{article['date'][4:6]}-{article['date'][6:]}",
            "press":            article["press"],
            "full_text":        full_text,
            "politicians_found": politicians_found,
            "ideologies_found": list(set(assembly[p]["ideology"] for p in politicians_found)),
        })

        time.sleep(0.5)

    # 저장
    output_path = "data/raw/articles_with_text_all.json"
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)

    print(f"\n✅ 완료")
    print(f"  - 크롤링 성공: {len(results)}건")
    print(f"  - 실패/URL없음: {failed}건")
    print(f"  - 저장: {output_path}")

    # 이념 분포
    from collections import Counter
    ideology_counter = Counter(
        ideology
        for a in results
        for ideology in a["ideologies_found"]
    )
    print(f"\n등장 정치인 이념 분포:")
    for k, v in ideology_counter.items():
        print(f"  {k}: {v}건")


if __name__ == "__main__":
    import sys
    # 엑셀 파일 경로를 인자로 받거나 기본값 사용
    excel_path = sys.argv[1] if len(sys.argv) > 1 else "data/raw/NewsResult_20240608-20260618.xlsx"
    crawl_from_bigkinds(excel_path)
